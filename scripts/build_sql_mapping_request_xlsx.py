#!/usr/bin/env python3
"""Generate War Room → 1C/SQL mapping request Excel for IT department.

Output: docs/war_room_sql_mapping_request.xlsx
Does not invent 1C technical object numbers. Does not embed secrets.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "docs" / "war_room_sql_mapping_request.xlsx"
HOME_COPY = Path.home() / "war_room_sql_mapping_request.xlsx"

# Styles
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
TITLE_FONT = Font(bold=True, name="Calibri", size=16, color="1F4E79")
SECTION_FONT = Font(bold=True, name="Calibri", size=12, color="1F4E79")
NORMAL_FONT = Font(name="Calibri", size=11)
YELLOW = PatternFill("solid", fgColor="FFF2CC")  # IT fill
RED = PatternFill("solid", fgColor="F4CCCC")  # mandatory without SQL
LIGHT_BLUE = PatternFill("solid", fgColor="D6EAF8")
LIGHT_GREEN = PatternFill("solid", fgColor="D5F5E3")
GREY = PatternFill("solid", fgColor="F2F3F4")
THIN = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)
WRAP = Alignment(wrap_text=True, vertical="top")

STATUS_IT = "Требуется заполнение ИТ"
STATUS_CHECK = "Не проверено"

# Registry columns
REG_COLS = [
    "№",
    "Раздел War Room",
    "Метрика",
    "Обязательность для первого SQL-релиза",
    "Бизнес-определение",
    "Формула расчёта в War Room",
    "Периодичность",
    "Гранулярность",
    "Требуемые разрезы",
    "Что должен вернуть SQL-запрос",
    "Объект 1С",
    "SQL-схема",
    "SQL-таблица / view",
    "SQL-поле / выражение",
    "Поле даты",
    "Поле магазина",
    "Фильтры / условия отбора",
    "Правила НДС",
    "Правила возвратов / сторно",
    "Правила закрытия периода",
    "Статус ИТ",
    "Ответственный ИТ",
    "Комментарий ИТ",
    "Статус сверки с Excel",
    "Комментарий по сверке",
]

# IT-fill columns (yellow) — indices 1-based in sheet after headers
IT_COLS = {
    "Объект 1С",
    "SQL-схема",
    "SQL-таблица / view",
    "SQL-поле / выражение",
    "Поле даты",
    "Поле магазина",
    "Фильтры / условия отбора",
    "Правила НДС",
    "Правила возвратов / сторно",
    "Правила закрытия периода",
    "Статус ИТ",
    "Ответственный ИТ",
    "Комментарий ИТ",
    "Статус сверки с Excel",
    "Комментарий по сверке",
}


def _metric(
    section: str,
    name: str,
    required: str,
    definition: str,
    formula: str,
    period: str,
    grain: str,
    dims: str,
    sql_returns: str,
) -> dict:
    return {
        "Раздел War Room": section,
        "Метрика": name,
        "Обязательность для первого SQL-релиза": required,
        "Бизнес-определение": definition,
        "Формула расчёта в War Room": formula,
        "Периодичность": period,
        "Гранулярность": grain,
        "Требуемые разрезы": dims,
        "Что должен вернуть SQL-запрос": sql_returns,
        "Объект 1С": "",
        "SQL-схема": "",
        "SQL-таблица / view": "",
        "SQL-поле / выражение": "",
        "Поле даты": "",
        "Поле магазина": "",
        "Фильтры / условия отбора": "",
        "Правила НДС": "",
        "Правила возвратов / сторно": "",
        "Правила закрытия периода": "",
        "Статус ИТ": STATUS_IT,
        "Ответственный ИТ": "",
        "Комментарий ИТ": "",
        "Статус сверки с Excel": STATUS_CHECK,
        "Комментарий по сверке": "",
    }


def build_metrics() -> list[dict]:
    S = "Технические измерения"
    P = "Продажи"
    C = "Чеки и покупатели"
    L = "P&L и маржинальность"
    I = "Запасы и выбытие"
    H = "Персонал"
    M = "Управленческие показатели"

    dim = "дата × магазин"
    dim_m = "магазин"
    dim_p = "дата × магазин × (опц. товар/категория)"

    rows: list[dict] = []

    # --- Technical dimensions ---
    rows += [
        _metric(S, "Дата операции", "Да", "Календарная дата хозяйственной операции / продажи / движения.", "Как в источнике 1С (дата документа/движения).", "День", "Дата", dim, "sale_date / operation_date"),
        _metric(S, "День", "Да", "Календарный день отчётности War Room.", "DATE(дата операции)", "День", "Дата", dim, "calendar_day"),
        _metric(S, "Неделя", "Да", "Отчётная неделя (нумерация и границы — по правилу компании).", "Неделя от даты операции (стандарт ISO или корпоративный — уточнить ИТ).", "Неделя", "Неделя", "неделя × магазин", "week_id / week_start"),
        _metric(S, "Месяц", "Да", "Отчётный месяц.", "YYYY-MM от даты операции", "Месяц", "Месяц", "месяц × магазин", "month_id"),
        _metric(S, "Код магазина", "Да", "Устойчивый код торговой точки в учётной системе.", "Справочник магазинов 1С", "Справочник", "Магазин", dim_m, "store_code"),
        _metric(S, "Наименование магазина", "Да", "Отображаемое имя магазина в War Room.", "Справочник магазинов 1С", "Справочник", "Магазин", dim_m, "store_name"),
        _metric(S, "Активность/статус магазина", "Да", "Признак, что магазин участвует в операционной отчётности (открыт/закрыт/пилот и т.п.).", "Статус из справочника / регистра сведений", "Справочник", "Магазин", dim_m, "store_status / is_active"),
        _metric(S, "Регион/кластер", "Нет", "Региональная или кластерная группировка магазинов, если ведётся в 1С.", "Справочник / доп. реквизиты магазина", "Справочник", "Регион/кластер", "регион × кластер × магазин", "region, cluster"),
    ]

    # --- Sales ---
    rows += [
        _metric(P, "Выручка факт", "Да", "Продажи розничного магазина за выбранный период — день, неделю, месяц или другой равнозначный период.", "Сумма розничных продаж за период и набор магазинов (правила НДС/возвратов — от ИТ).", "День / неделя / месяц", dim, dim, "revenue_fact"),
        _metric(P, "Выручка план", "Да", "Плановая выручка магазина (сети) на тот же период, что и факт.", "Значение плана из регистра/документа планирования 1С.", "День / неделя / месяц", dim, dim, "revenue_plan"),
        _metric(P, "Отклонение факта выручки от плана, руб.", "Да", "Абсолютное отклонение факта от плана.", "выручка_факт − выручка_план", "День / неделя / месяц", dim, dim, "revenue_fact - revenue_plan"),
        _metric(P, "Выполнение плана выручки, %", "Да", "Относительное выполнение плана продаж.", "выручка_факт / выручка_план × 100%", "День / неделя / месяц", dim, dim, "revenue_fact / revenue_plan * 100"),
        _metric(P, "Выручка прошлого периода / прошлого года", "Нет", "Сопоставимая выручка за аналогичный прошлый период (WoW/YoY), если доступна в 1С.", "Выручка за сопоставимый период по правилу компании.", "День / неделя / месяц", dim, dim, "revenue_ly / revenue_prev"),
        _metric(P, "Возвраты", "Да", "Сумма возвратов покупателей, уменьшающая или отдельно отражаемая относительно валовой продажи.", "Сумма возвратных операций (правило знака — от ИТ).", "День / неделя / месяц", dim, dim, "returns_amount"),
        _metric(P, "Сторно", "Да", "Сторнирующие движения/документы, влияющие на продажи.", "Сумма сторно по правилам проведения 1С.", "День / неделя / месяц", dim, dim, "storno_amount"),
        _metric(P, "Скидки", "Нет", "Сумма предоставленных скидок за период.", "Сумма скидок из чеков/продаж.", "День / неделя / месяц", dim, dim, "discount_amount"),
        _metric(P, "Продажи до скидок", "Нет", "Выручка до применения скидок (если ведётся).", "Сумма до скидки / цена × количество.", "День / неделя / месяц", dim, dim, "gross_sales_before_discount"),
    ]

    # --- Checks ---
    rows += [
        _metric(C, "Количество чеков", "Да", "Количество чеков из отчёта о продажах ККМ за выбранный период.", "COUNTDistinct чеков / показатель из отчёта ККМ (уточнить ИТ).", "День / неделя / месяц", dim, dim, "receipt_count"),
        _metric(C, "Средний чек", "Да", "Средний размер покупки на один чек за одинаковый период и одинаковый набор магазинов.", "выручка / количество чеков", "День / неделя / месяц", dim, dim, "revenue_fact / receipt_count"),
        _metric(C, "Количество проданных единиц товара", "Нет", "Суммарное количество проданных единиц (SKU/упаковок — уточнить единицу).", "SUM(количество)", "День / неделя / месяц", dim_p, dim_p, "item_quantity"),
        _metric(C, "Товаров в чеке", "Нет", "Среднее число позиций/единиц на один чек.", "item_quantity / receipt_count", "День / неделя / месяц", dim, dim, "item_quantity / receipt_count"),
        _metric(C, "Количество клиентов", "Нет", "Число уникальных клиентов/покупателей, если показатель ведётся в системе.", "COUNT DISTINCT клиентов (если есть идентификатор).", "День / неделя / месяц", dim, dim, "customer_count"),
    ]

    # --- P&L ---
    rows += [
        _metric(L, "Себестоимость продаж", "Да", "Себестоимость товаров по правилам партионного учёта, применяемого в 1С.", "Сумма себестоимости реализованного товара (метод партий — подтвердить ИТ).", "День / неделя / месяц", dim, dim, "cogs"),
        _metric(L, "Валовая прибыль", "Да", "Прибыль после себестоимости до операционных расходов.", "выручка − себестоимость", "День / неделя / месяц", dim, dim, "revenue_fact - cogs"),
        _metric(L, "Валовая маржа, %", "Да", "Отношение валовой прибыли к выручке.", "валовая_прибыль / выручка × 100%", "День / неделя / месяц", dim, dim, "(revenue_fact - cogs) / revenue_fact * 100"),
        _metric(L, "План валовой прибыли", "Нет", "Плановая валовая прибыль, если утверждается в 1С.", "Значение плана.", "Месяц", dim, dim, "gross_profit_plan"),
        _metric(L, "План валовой маржи, %", "Нет", "Плановая валовая маржа, если утверждается.", "План ВП / план выручки × 100% или готовое поле.", "Месяц", dim, dim, "gross_margin_plan_pct"),
        _metric(L, "OPEX всего", "Да", "Сумма операционных расходов магазина/сети за период.", "SUM(статей OPEX) по утверждённому перечню.", "Месяц", "месяц × магазин", "месяц × магазин × статья", "opex_total"),
        _metric(L, "ФОТ", "Да", "Фонд оплаты труда (начисления), относимый к магазину/периоду.", "Сумма ФОТ по правилам учёта персонала.", "Месяц", "месяц × магазин", "месяц × магазин", "payroll"),
        _metric(L, "Аренда", "Нет", "Расходы на аренду торговых площадей.", "Сумма по статье «Аренда».", "Месяц", "месяц × магазин", "месяц × магазин", "rent"),
        _metric(L, "Коммунальные услуги", "Нет", "Коммунальные и энергозатраты.", "Сумма по статье.", "Месяц", "месяц × магазин", "месяц × магазин", "utilities"),
        _metric(L, "Маркетинг", "Нет", "Маркетинговые и рекламные расходы.", "Сумма по статье.", "Месяц", "месяц × магазин", "месяц × магазин", "marketing"),
        _metric(L, "Эквайринг / банковские комиссии", "Нет", "Комиссии банков и эквайринга.", "Сумма по статье.", "Месяц", "месяц × магазин", "месяц × магазин", "acquiring_fees"),
        _metric(L, "Логистика", "Нет", "Логистические расходы, относимые на магазин.", "Сумма по статье.", "Месяц", "месяц × магазин", "месяц × магазин", "logistics"),
        _metric(L, "Ремонт и эксплуатация", "Нет", "Ремонт, содержание, эксплуатация.", "Сумма по статье.", "Месяц", "месяц × магазин", "месяц × магазин", "maintenance"),
        _metric(L, "IT / связь", "Нет", "IT, связь, ПО, сервисы.", "Сумма по статье.", "Месяц", "месяц × магазин", "месяц × магазин", "it_telecom"),
        _metric(L, "Прочие операционные расходы", "Нет", "Прочие OPEX, не попавшие в перечисленные статьи.", "Сумма прочих статей.", "Месяц", "месяц × магазин", "месяц × магазин", "other_opex"),
        _metric(L, "EBITDA / операционная прибыль", "Да", "Операционный финансовый итог магазина/сети. Точная формула должна быть подтверждена ИТ и финансами.", "Требуется подтверждение ИТ/финансов (типично: валовая прибыль − OPEX; уточнить).", "Месяц", "месяц × магазин", "месяц × магазин", "ebitda / operating_profit"),
    ]

    # --- Inventory ---
    rows += [
        _metric(I, "Товарный остаток в штуках на дату", "Да", "Остаток товара на конец выбранной даты.", "Остаток на конец дня из регистра накопления остатков.", "День (снимок)", "дата × магазин × (товар)", "дата × магазин × товар", "stock_quantity"),
        _metric(I, "Товарный остаток в рублях на дату", "Да", "Стоимостная оценка остатка на конец даты.", "Остаток в сумме по правилам оценки запасов 1С.", "День (снимок)", "дата × магазин × (товар)", "дата × магазин × товар", "stock_value"),
        _metric(I, "Списание товара", "Да", "Выбытие товара с обязательным разделением по подтипам списания.", "Сумма/количество списаний по типам.", "День / неделя / месяц", dim_p, "дата × магазин × тип/подтип", "writeoff_amount / writeoff_qty"),
        _metric(I, "Подтип списания", "Да", "Классификация причины/вида списания (ФРОВ, брак, срок годности и т.п. — перечень от ИТ).", "Классификатор вида выбытия.", "Справочник / движение", "тип × подтип", "тип × подтип", "writeoff_subtype"),
        _metric(I, "Сумма списания", "Да", "Стоимость списанного товара.", "SUM(сумма списания)", "День / неделя / месяц", dim, dim, "writeoff_amount"),
        _metric(I, "Количество списанного товара", "Да", "Количество списанных единиц.", "SUM(количество)", "День / неделя / месяц", dim, dim, "writeoff_quantity"),
        _metric(I, "Недостача", "Да", "Выявленная недостача (инвентаризация / учёт).", "Сумма/количество недостач.", "День / месяц", dim, dim, "shortage_amount / shortage_qty"),
        _metric(I, "Перемещение товара", "Нет", "Внутренние перемещения между складами/магазинами.", "Движения перемещения (приход/расход).", "День", "дата × магазин-отправитель/получатель", "from_store × to_store", "transfer_qty / transfer_amount"),
        _metric(I, "Инвентаризационная корректировка", "Нет", "Корректировки остатков по результатам инвентаризации.", "Документы инвентаризации / корректировки.", "День / месяц", dim, dim, "inventory_adjustment"),
        _metric(I, "Возврат поставщику", "Нет", "Выбытие товара возвратом поставщику.", "Документы возврата поставщику.", "День / месяц", dim, dim, "return_to_supplier"),
        _metric(I, "Продажа/выбытие товара", "Нет", "Выбытие товара через продажу (для связки с остатками).", "Количество/сумма реализованного.", "День", dim_p, dim_p, "sold_quantity"),
        _metric(I, "Оборачиваемость запасов", "Нет", "Показатель оборачиваемости (если требуется в War Room).", "Формула компании (например себестоимость / средний остаток) — подтвердить.", "Месяц", "месяц × магазин", "месяц × магазин", "inventory_turnover"),
        _metric(I, "Out of Stock / отсутствие товара", "Нет", "Признак/доля отсутствующих позиций (если доступно).", "Правило расчёта OOS от ИТ/категорийного управления.", "День / неделя", "дата × магазин × (категория)", "дата × магазин", "oos_pct / oos_flag"),
    ]

    # --- Staff ---
    rows += [
        _metric(H, "Штатная численность", "Нет", "Число штатных единиц по магазину.", "Штатное расписание / регистр кадров.", "Месяц", "месяц × магазин", "месяц × магазин", "headcount_plan"),
        _metric(H, "Фактическая численность", "Нет", "Фактически занятые сотрудники.", "Списочная / явочная численность — уточнить.", "Месяц", "месяц × магазин", "месяц × магазин", "headcount_fact"),
        _metric(H, "Вакансии", "Нет", "Разница штат − факт или отдельный учёт вакансий.", "штат − факт (или готовое поле).", "Месяц", "месяц × магазин", "месяц × магазин", "vacancies"),
        _metric(H, "ФОТ (персонал)", "Да", "ФОТ в контуре персонала (может совпадать со статьёй P&L).", "Сумма начислений за период.", "Месяц", "месяц × магазин", "месяц × магазин", "payroll"),
        _metric(H, "Производительность на сотрудника", "Нет", "Операционная производительность (чеки/единицы на сотрудника).", "чеки или единицы / численность", "Месяц", "месяц × магазин", "месяц × магазин", "receipt_count / headcount_fact"),
        _metric(H, "Выручка на сотрудника", "Нет", "Выручка, приходящаяся на одного сотрудника.", "выручка / численность", "Месяц", "месяц × магазин", "месяц × магазин", "revenue_fact / headcount_fact"),
        _metric(H, "ФОТ как процент выручки", "Нет", "Доля ФОТ в выручке.", "ФОТ / выручка × 100%", "Месяц", "месяц × магазин", "месяц × магазин", "payroll / revenue_fact * 100"),
    ]

    # --- Management ---
    rows += [
        _metric(M, "Красный/жёлтый/зелёный статус KPI", "Нет", "Светофор статуса KPI по порогам War Room.", "Расчёт на стороне War Room по порогам; SQL отдаёт исходные метрики.", "День / неделя / месяц", dim, dim, "базовые KPI-поля"),
        _metric(M, "Индекс критичности магазина", "Нет", "Сводный индекс риска магазина.", "Расчёт War Room на основе KPI; SQL — входные показатели.", "День / неделя / месяц", dim_m, dim_m, "набор KPI"),
        _metric(M, "Рейтинг проблемных магазинов", "Нет", "Ранжирование магазинов по отклонениям.", "Расчёт War Room; SQL — факты.", "День / неделя / месяц", dim_m, dim_m, "набор KPI"),
        _metric(M, "План/факт по ключевым показателям", "Да", "Сопоставление плана и факта по ключевым KPI релиза.", "Пары plan/fact из SQL views.", "День / неделя / месяц", dim, dim, "plan_* и fact_*"),
        _metric(M, "Дата последнего обновления данных", "Да", "Метка свежести данных для UI War Room.", "MAX(updated_at) по используемым views/таблицам.", "Онлайн / по обновлению", "источник", "—", "updated_at"),
    ]

    return rows


def style_header(ws, row: int, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = THIN


def autosize(ws, widths: dict[int, int] | None = None, max_width: int = 42) -> None:
    if widths:
        for col, w in widths.items():
            ws.column_dimensions[get_column_letter(col)].width = w
        return
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        maxlen = 10
        for row in range(1, min(ws.max_row, 80) + 1):
            val = ws.cell(row=row, column=col).value
            if val is not None:
                maxlen = max(maxlen, min(len(str(val)), max_width))
        ws.column_dimensions[letter].width = maxlen + 2


def freeze_filter(ws, header_row: int = 1) -> None:
    ws.freeze_panes = f"A{header_row + 1}"
    if ws.max_row >= header_row and ws.max_column >= 1:
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(ws.max_column)}{ws.max_row}"


def add_table(ws, name: str, header_row: int = 1) -> None:
    ref = f"A{header_row}:{get_column_letter(ws.max_column)}{ws.max_row}"
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(table)


def sheet_instruction(wb: Workbook, metrics: list[dict]) -> None:
    ws = wb.create_sheet("Инструкция для ИТ", 0)
    ws["A1"] = "War Room — задание на сопоставление метрик с объектами 1С / SQL"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:B1")

    mandatory = sum(1 for m in metrics if m["Обязательность для первого SQL-релиза"] == "Да")
    optional = len(metrics) - mandatory

    lines = [
        ("", ""),
        ("Цель документа", "Определить, из каких объектов 1С и таблиц/views SQL должны поступать данные в War Room."),
        ("", ""),
        ("Контекст", ""),
        ("•", "War Room подключается к Microsoft SQL Server, база retail (сервер уже известен ИТ-отделу заказчика)."),
        ("•", "SQL-пользователь используется только на чтение (SELECT). Изменяющие операции запрещены."),
        ("•", "База retail является SQL-копией 1С:Предприятие с техническими именами (_AccumRg*, _Reference*, _Document*, _Fld*)."),
        ("•", "В этом документе нельзя указывать пароли, строки подключения, токены и другие секреты."),
        ("•", "ИТ заполняет реальные объекты 1С, SQL-таблицы/views, поля, фильтры и бизнес-правила. Не заполненные жёлтые ячейки означают «требуется ответ ИТ»."),
        ("•", "Предпочтительно создать SQL views с понятными бизнес-названиями, а не подключать War Room напрямую к техническим таблицам _AccumRg#### / _Fld####."),
        ("•", "Все расчёты должны быть подтверждены на контрольном периоде сверкой с текущим Excel-отчётом War Room (лист «Сверка Excel и SQL»)."),
        ("", ""),
        ("Краткая сводка готовности SQL", ""),
        ("Всего метрик в реестре", str(len(metrics))),
        ("Обязательных для первого SQL-релиза", str(mandatory)),
        ("Необязательных (второй этап)", str(optional)),
        ("Метрик с подтверждённым SQL-источником", "0 (все ожидают заполнения ИТ)"),
        ("Готовность первого SQL-релиза", "0% — требуется карта объектов 1С → SQL и сверка"),
        ("Рекомендуемый путь", "SQL views (лист «Требования к SQL views») + заполнение реестра + контрольная сверка"),
        ("", ""),
        ("Что требуется от ИТ", ""),
        ("1", "Подтвердить, что база retail является корректным источником данных для War Room."),
        ("2", "Указать объект 1С и SQL-таблицу/регистр/view для каждой метрики (особенно обязательных)."),
        ("3", "Указать реальные поля SQL (или выражения во view)."),
        ("4", "Указать, каким полем/справочником определяется магазин."),
        ("5", "Указать, как определяются дата, неделя, месяц и отчётный период (границы недели/месяца)."),
        ("6", "Указать правила НДС, возвратов, сторно, отменённых документов и закрытия периода."),
        ("7", "Указать типы и подтипы списания / выбытия товаров."),
        ("8", "Указать, как рассчитывается себестоимость в партионном учёте."),
        ("9", "Подтвердить выборку цифр за один согласованный контрольный период для сверки с Excel."),
        ("10", "По возможности создать views с понятными именами для War Room (см. лист «Требования к SQL views»)."),
        ("", ""),
        ("Как работать с файлом", ""),
        ("•", "Лист «Реестр метрик War Room» — основной: жёлтые колонки заполняет ИТ; красным выделены обязательные метрики без SQL-источника."),
        ("•", "Лист «Карточка SQL-объекта 1С» — детализация каждого указанного объекта."),
        ("•", "Лист «Требования к SQL views» — целевая витрина для приложения."),
        ("•", "Лист «Сверка Excel и SQL» — контрольные цифры; статус считается по разнице."),
        ("•", "Лист «Справочник статусов» — значения для выпадающих списков."),
        ("", ""),
        ("Ограничения", ""),
        ("•", "Не угадывать соответствие _AccumRg#### / _Fld#### бизнес-метрике без подтверждения."),
        ("•", "Не менять пароли и не публиковать секреты в этом файле или переписке."),
        ("•", "War Room не объявляет SQL-источник готовым, пока не закрыта сверка обязательных метрик."),
    ]

    r = 3
    for a, b in lines:
        ws.cell(row=r, column=1, value=a).font = SECTION_FONT if a and not a.startswith("•") and not a.isdigit() and a not in {"•"} else NORMAL_FONT
        cell_b = ws.cell(row=r, column=2, value=b)
        cell_b.font = NORMAL_FONT
        cell_b.alignment = WRAP
        if a == "Краткая сводка готовности SQL":
            ws.cell(row=r, column=1).fill = LIGHT_BLUE
            cell_b.fill = LIGHT_BLUE
        if a in {"Всего метрик в реестре", "Обязательных для первого SQL-релиза", "Метрик с подтверждённым SQL-источником", "Готовность первого SQL-релиза"}:
            ws.cell(row=r, column=1).fill = GREY
            cell_b.fill = GREY
        if a == "Готовность первого SQL-релиза":
            cell_b.fill = RED
        r += 1

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 110
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A3"


def sheet_registry(wb: Workbook, metrics: list[dict]) -> None:
    ws = wb.create_sheet("Реестр метрик War Room", 1)
    for c, name in enumerate(REG_COLS, 1):
        ws.cell(row=1, column=c, value=name)
    style_header(ws, 1, len(REG_COLS))
    ws.row_dimensions[1].height = 45

    it_col_idx = {name: i for i, name in enumerate(REG_COLS, 1)}

    for i, m in enumerate(metrics, 1):
        row = i + 1
        values = [i] + [m[c] for c in REG_COLS[1:]]
        for c, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.font = NORMAL_FONT
            cell.alignment = WRAP
            cell.border = THIN
            col_name = REG_COLS[c - 1]
            if col_name in IT_COLS:
                cell.fill = YELLOW
            if col_name == "Обязательность для первого SQL-релиза" and val == "Да":
                # mark metric name red — mandatory without confirmed SQL
                name_cell = ws.cell(row=row, column=it_col_idx["Метрика"])
                name_cell.fill = RED
                req_cell = ws.cell(row=row, column=it_col_idx["Обязательность для первого SQL-релиза"])
                req_cell.fill = RED

    # Data validations
    dv_it = DataValidation(
        type="list",
        formula1='"Требуется заполнение ИТ,В работе,Найдено,Подтверждено,Нет в базе,Не требуется"',
        allow_blank=True,
    )
    dv_check = DataValidation(
        type="list",
        formula1='"Не проверено,Совпало,Расхождение,Ожидаемое различие"',
        allow_blank=True,
    )
    dv_req = DataValidation(type="list", formula1='"Да,Нет"', allow_blank=False)
    ws.add_data_validation(dv_it)
    ws.add_data_validation(dv_check)
    ws.add_data_validation(dv_req)
    last = ws.max_row
    dv_it.add(f"{get_column_letter(it_col_idx['Статус ИТ'])}2:{get_column_letter(it_col_idx['Статус ИТ'])}{last}")
    dv_check.add(f"{get_column_letter(it_col_idx['Статус сверки с Excel'])}2:{get_column_letter(it_col_idx['Статус сверки с Excel'])}{last}")
    dv_req.add(f"{get_column_letter(it_col_idx['Обязательность для первого SQL-релиза'])}2:{get_column_letter(it_col_idx['Обязательность для первого SQL-релиза'])}{last}")

    freeze_filter(ws, 1)
    add_table(ws, "RegistryMetrics")
    autosize(
        ws,
        {
            1: 5,
            2: 24,
            3: 36,
            4: 14,
            5: 40,
            6: 32,
            7: 16,
            8: 22,
            9: 28,
            10: 28,
            11: 22,
            12: 12,
            13: 24,
            14: 24,
            15: 16,
            16: 16,
            17: 24,
            18: 18,
            19: 22,
            20: 22,
            21: 22,
            22: 18,
            23: 24,
            24: 18,
            25: 24,
        },
    )


def sheet_cards(wb: Workbook) -> None:
    ws = wb.create_sheet("Карточка SQL-объекта 1С", 2)
    cols = [
        "№",
        "Бизнес-название объекта",
        "Тип объекта 1С",
        "Имя объекта 1С",
        "Техническая SQL-таблица",
        "Роль объекта",
        "Поле даты",
        "Поле магазина",
        "Поле суммы",
        "Поле количества",
        "Поле себестоимости",
        "Поле документа-основания",
        "Поле статуса проведения",
        "Поле типа операции",
        "Нужные соединения с другими таблицами",
        "Условия исключения",
        "Пример SQL SELECT",
        "Подтверждено ИТ",
        "Комментарий ИТ",
    ]
    for c, name in enumerate(cols, 1):
        ws.cell(row=1, column=c, value=name)
    style_header(ws, 1, len(cols))

    examples = [
        ("Регистр продаж", "Регистр накопления", "Источник выручки / чеков / количества", "Продажи розницы за период"),
        ("Регистр себестоимости", "Регистр накопления", "Источник COGS по партионному учёту", "Себестоимость реализованного"),
        ("Регистр остатков", "Регистр накопления", "Снимки товарных остатков", "Остатки на дату"),
        ("Документ списания", "Документ", "Выбытие товара по типам/подтипам", "Списания и потери"),
        ("Справочник магазинов", "Справочник", "Код, наименование, статус, регион/кластер", "Измерение магазина"),
        ("Регистр планов", "Регистр сведений", "Планы выручки / маржи / прочих KPI", "План продаж"),
        ("Регистр расходов (OPEX)", "Регистр накопления / сведений", "Операционные статьи расходов", "OPEX / ФОТ / аренда и др."),
        ("Кадровый регистр / ФОТ", "Регистр сведений / накопления", "Численность и начисления", "Персонал"),
    ]

    yellow_cols = set(range(3, 20))  # IT fills almost everything except № and business name partially
    for i, (biz, typ, role, comment) in enumerate(examples, 1):
        row = i + 1
        data = [
            i,
            biz,
            typ,
            "",  # Имя объекта 1С — НЕ выдумываем
            "",  # Техническая SQL-таблица — НЕ выдумываем номера
            role,
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "Исключить непроведённые / помеченные на удаление — уточнить ИТ",
            "SELECT ... — заполнить после указания реальных таблиц/views",
            "Нет",
            f"{comment}. Технические имена (_AccumRg#### / _Fld####) — только от ИТ.",
        ]
        for c, val in enumerate(data, 1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.font = NORMAL_FONT
            cell.alignment = WRAP
            cell.border = THIN
            if c in yellow_cols or c in {3, 4, 5, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19}:
                cell.fill = YELLOW
            if c == 2:
                cell.fill = LIGHT_BLUE

    dv_type = DataValidation(
        type="list",
        formula1='"Регистр накопления,Регистр сведений,Документ,Справочник,План,View,Другое"',
        allow_blank=True,
    )
    dv_yes = DataValidation(type="list", formula1='"Да,Нет,Частично"', allow_blank=True)
    ws.add_data_validation(dv_type)
    ws.add_data_validation(dv_yes)
    dv_type.add(f"C2:C{ws.max_row}")
    dv_yes.add(f"R2:R{ws.max_row}")

    freeze_filter(ws, 1)
    add_table(ws, "ObjectCards")
    for col in range(1, len(cols) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 22 if col > 1 else 5
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["F"].width = 36
    ws.column_dimensions["Q"].width = 40
    ws.column_dimensions["S"].width = 40


def sheet_views(wb: Workbook) -> None:
    ws = wb.create_sheet("Требования к SQL views", 3)
    cols = [
        "№",
        "Имя view",
        "Назначение",
        "Требуемые колонки",
        "Минимальная гранулярность",
        "Ключевые фильтры",
        "Что не должно попадать в данные",
        "Кто подтверждает цифры",
        "Статус ИТ",
        "Комментарий ИТ",
    ]
    for c, name in enumerate(cols, 1):
        ws.cell(row=1, column=c, value=name)
    style_header(ws, 1, len(cols))

    views = [
        (
            "vw_warroom_sales_daily",
            "Ежедневные продажи по магазинам для KPI выручки, чеков, возвратов и скидок.",
            "sale_date; store_id; store_code; store_name; revenue_fact; receipt_count; item_quantity; returns_amount; discount_amount; net_revenue; vat_flag (или эквивалент правила НДС); updated_at",
            "день × магазин",
            "Период; магазины; только проведённые операции; согласованный контур розницы",
            "Непроведённые документы; помеченные на удаление; служебные/тестовые точки (если есть); операции вне розничного контура — уточнить ИТ",
            "ИТ + финансы/ритейл-аналитика",
        ),
        (
            "vw_warroom_sales_plan",
            "Планы продаж для сравнения план/факт.",
            "period_start; period_grain (day/week/month); store_id; store_code; store_name; revenue_plan; gross_profit_plan (опц.); updated_at",
            "период × магазин",
            "Актуальная версия плана; утверждённые планы",
            "Черновики планов; устаревшие версии без пометки",
            "ИТ + финансы",
        ),
        (
            "vw_warroom_store_directory",
            "Справочник магазинов и статусов.",
            "store_id; store_code; store_name; is_active; region; cluster; format; opened_at; closed_at; updated_at",
            "магазин",
            "Актуальные элементы справочника",
            "Помеченные на удаление без необходимости; технические склады, если не входят в отчётность",
            "ИТ + операционный блок",
        ),
        (
            "vw_warroom_cogs_daily",
            "Себестоимость продаж по правилам партионного учёта.",
            "sale_date; store_id; store_code; cogs; revenue_fact (опц. для контроля); updated_at",
            "день × магазин",
            "Согласованный метод партий; закрытые периоды — по правилу",
            "Движения без проведения; тестовые базы",
            "ИТ + финансы",
        ),
        (
            "vw_warroom_opex_monthly",
            "Операционные расходы по статьям (OPEX, ФОТ, аренда и др.).",
            "month_id; store_id; store_code; expense_article; amount; payroll; rent; utilities; marketing; acquiring_fees; logistics; maintenance; it_telecom; other_opex; opex_total; updated_at",
            "месяц × магазин × статья",
            "Утверждённые статьи управленческого учёта",
            "Капитальные затраты, если не входят в OPEX; внутригрупповые обороты — по правилу",
            "ИТ + финансы",
        ),
        (
            "vw_warroom_inventory_daily",
            "Товарные остатки на конец дня.",
            "snapshot_date; store_id; product_id; product_name; stock_quantity; stock_value; updated_at",
            "дата × магазин × товар",
            "Снимок на конец дня; согласованная оценка стоимости",
            "Нулевые/служебные номенклатуры по согласованию; закрытые склады вне периметра",
            "ИТ + категорийный/операционный блок",
        ),
        (
            "vw_warroom_writeoffs",
            "Списания и выбытия с типами/подтипами.",
            "operation_date; store_id; product_id; writeoff_type; writeoff_subtype; quantity; amount; document_number; updated_at",
            "дата × магазин × документ/строка",
            "Проведённые документы списания; классификатор подтипов",
            "Непроведённые; отменённые; перемещения, если они не относятся к списанию",
            "ИТ + операционный блок",
        ),
        (
            "vw_warroom_headcount",
            "Численность и ФОТ для блока персонала.",
            "month_id; store_id; headcount_plan; headcount_fact; vacancies; payroll; updated_at",
            "месяц × магазин",
            "Согласованный контур сотрудников магазина",
            "Сотрудники бэкофиса, если не относятся к магазину — уточнить",
            "ИТ + HR/финансы",
        ),
        (
            "vw_warroom_kpi_monthly",
            "Витрина месячных KPI для управленческих экранов.",
            "month_id; store_id; store_code; store_name; revenue_fact; revenue_plan; receipt_count; avg_ticket; cogs; gross_profit; gross_margin_pct; opex_total; ebitda; writeoff_amount; stock_value_eom; updated_at",
            "месяц × магазин",
            "Закрытый/оперативный месяц — по правилу компании",
            "Несогласованные черновые оценки без флага",
            "ИТ + финансы + владелец War Room",
        ),
    ]

    for i, rowdata in enumerate(views, 1):
        row = i + 1
        values = [i, *rowdata, STATUS_IT, ""]
        for c, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.font = NORMAL_FONT
            cell.alignment = WRAP
            cell.border = THIN
            if c >= 9:
                cell.fill = YELLOW
            if c == 2:
                cell.fill = LIGHT_GREEN

    dv_it = DataValidation(
        type="list",
        formula1='"Требуется заполнение ИТ,В работе,Найдено,Подтверждено,Нет в базе,Не требуется"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_it)
    dv_it.add(f"I2:I{ws.max_row}")

    note_row = ws.max_row + 2
    ws.cell(row=note_row, column=1, value="Примечание").font = SECTION_FONT
    ws.cell(
        row=note_row + 1,
        column=1,
        value=(
            "Views должны скрывать технические имена 1С (_AccumRg*, _Fld*). "
            "War Room подключается к views, а не напрямую к сырым таблицам, если ИТ может их предоставить. "
            "Пример SQL SELECT и точные источники заполняются ИТ после карты объектов."
        ),
    ).alignment = WRAP
    ws.merge_cells(start_row=note_row + 1, start_column=1, end_row=note_row + 1, end_column=10)

    freeze_filter(ws, 1)
    add_table(ws, "SqlViews")
    widths = {1: 5, 2: 28, 3: 36, 4: 55, 5: 22, 6: 32, 7: 36, 8: 24, 9: 24, 10: 24}
    autosize(ws, widths)


def sheet_reconciliation(wb: Workbook, metrics: list[dict]) -> None:
    ws = wb.create_sheet("Сверка Excel и SQL", 4)
    cols = [
        "№",
        "Метрика",
        "Период",
        "Магазин / фильтр",
        "Значение в Excel",
        "Значение в SQL",
        "Разница",
        "Допустимое отклонение",
        "Статус",
        "Причина расхождения",
        "Решение",
        "Ответственный",
        "Дата проверки",
    ]
    for c, name in enumerate(cols, 1):
        ws.cell(row=1, column=c, value=name)
    style_header(ws, 1, len(cols))

    # Pre-populate mandatory metrics as reconciliation checklist rows
    mandatory = [m["Метрика"] for m in metrics if m["Обязательность для первого SQL-релиза"] == "Да"]
    # Keep a practical subset for first pass + leave extra blank rows
    seed = mandatory[:25]
    for i, name in enumerate(seed, 1):
        row = i + 1
        ws.cell(row=row, column=1, value=i).border = THIN
        ws.cell(row=row, column=2, value=name).border = THIN
        ws.cell(row=row, column=2).fill = RED
        for c in range(3, 14):
            cell = ws.cell(row=row, column=c, value="")
            cell.border = THIN
            cell.fill = YELLOW
        # Default period / filter placeholders
        ws.cell(row=row, column=3, value="Указать контрольный период")
        ws.cell(row=row, column=4, value="Сеть / указать магазин")
        ws.cell(row=row, column=5, value=None)  # Excel value — numeric later
        ws.cell(row=row, column=6, value=None)
        # Difference formula
        ws.cell(row=row, column=7, value=f"=IF(OR(E{row}=\"\",F{row}=\"\"),\"\",F{row}-E{row})")
        ws.cell(row=row, column=8, value=0)
        # Status formula
        ws.cell(
            row=row,
            column=9,
            value=f'=IF(OR(E{row}="",F{row}=""),"Не проверено",IF(ABS(G{row})<=H{row},"Совпало","Расхождение"))',
        )

    # Extra empty rows for IT
    start = len(seed) + 2
    for i in range(start, start + 15):
        n = i - 1
        ws.cell(row=i, column=1, value=n).border = THIN
        for c in range(2, 14):
            cell = ws.cell(row=i, column=c, value="")
            cell.border = THIN
            cell.fill = YELLOW
        ws.cell(row=i, column=7, value=f"=IF(OR(E{i}=\"\",F{i}=\"\"),\"\",F{i}-E{i})")
        ws.cell(row=i, column=8, value=0)
        ws.cell(
            row=i,
            column=9,
            value=f'=IF(OR(E{i}="",F{i}=""),"Не проверено",IF(ABS(G{i})<=H{i},"Совпало","Расхождение"))',
        )

    # Conditional formatting for status
    red_fill = PatternFill("solid", fgColor="F4CCCC")
    green_fill = PatternFill("solid", fgColor="D5F5E3")
    ws.conditional_formatting.add(
        f"I2:I{ws.max_row}",
        FormulaRule(formula=['I2="Расхождение"'], fill=red_fill),
    )
    ws.conditional_formatting.add(
        f"I2:I{ws.max_row}",
        FormulaRule(formula=['I2="Совпало"'], fill=green_fill),
    )

    note = ws.max_row + 2
    ws.cell(row=note, column=1, value="Правила сверки").font = SECTION_FONT
    ws.cell(
        row=note + 1,
        column=1,
        value=(
            "Разница = Значение в SQL − Значение в Excel. "
            "Если |Разница| ≤ Допустимое отклонение → «Совпало», иначе → «Расхождение». "
            "Контрольный период и перечень магазинов согласовать до начала сверки. "
            "SQL-релиз не считается готовым при открытых расхождениях по обязательным метрикам."
        ),
    ).alignment = WRAP
    ws.merge_cells(start_row=note + 1, start_column=1, end_row=note + 1, end_column=13)

    freeze_filter(ws, 1)
    # Don't use Table here — Excel tables + many formulas can be finicky; auto filter is enough
    autosize(
        ws,
        {1: 5, 2: 36, 3: 22, 4: 22, 5: 16, 6: 16, 7: 12, 8: 16, 9: 14, 10: 24, 11: 20, 12: 16, 13: 14},
    )


def sheet_statuses(wb: Workbook) -> None:
    ws = wb.create_sheet("Справочник статусов", 5)
    ws["A1"] = "Справочник значений для выпадающих списков"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:B1")

    blocks = [
        (
            "Статус ИТ",
            [
                "Требуется заполнение ИТ",
                "В работе",
                "Найдено",
                "Подтверждено",
                "Нет в базе",
                "Не требуется",
            ],
        ),
        (
            "Статус сверки",
            ["Не проверено", "Совпало", "Расхождение", "Ожидаемое различие"],
        ),
        ("Обязательность", ["Да", "Нет"]),
        (
            "Тип объекта 1С",
            [
                "Регистр накопления",
                "Регистр сведений",
                "Документ",
                "Справочник",
                "План",
                "View",
                "Другое",
            ],
        ),
    ]

    col = 1
    for title, values in blocks:
        ws.cell(row=3, column=col, value=title).font = HEADER_FONT
        ws.cell(row=3, column=col).fill = HEADER_FILL
        for i, v in enumerate(values, 4):
            cell = ws.cell(row=i, column=col, value=v)
            cell.border = THIN
            cell.fill = GREY
        ws.column_dimensions[get_column_letter(col)].width = 28
        col += 1

    ws["A12"] = "Легенда оформления"
    ws["A12"].font = SECTION_FONT
    ws["A13"] = "Жёлтый"
    ws["B13"] = "Поле заполняет ИТ"
    ws["A13"].fill = YELLOW
    ws["A14"] = "Красный"
    ws["B14"] = "Обязательная метрика первого SQL-релиза без подтверждённого источника"
    ws["A14"].fill = RED
    ws["A15"] = "Голубой"
    ws["B15"] = "Справочная / контекстная информация"
    ws["A15"].fill = LIGHT_BLUE
    ws["A16"] = "Зелёный"
    ws["B16"] = "Рекомендуемые SQL views / успешная сверка"
    ws["A16"].fill = LIGHT_GREEN
    ws.freeze_panes = "A3"
    ws.column_dimensions["B"].width = 70


def main() -> None:
    metrics = build_metrics()
    wb = Workbook()
    # remove default
    default = wb.active
    wb.remove(default)

    sheet_instruction(wb, metrics)
    sheet_registry(wb, metrics)
    sheet_cards(wb)
    sheet_views(wb)
    sheet_reconciliation(wb, metrics)
    sheet_statuses(wb)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    wb.save(HOME_COPY)

    # verify open
    from openpyxl import load_workbook

    wb2 = load_workbook(OUT, data_only=False)
    sheets = wb2.sheetnames
    reg = wb2["Реестр метрик War Room"]
    # count metrics = rows - 1
    metric_count = reg.max_row - 1
    mandatory = 0
    for r in range(2, reg.max_row + 1):
        if reg.cell(row=r, column=4).value == "Да":
            mandatory += 1

    print("OK", OUT)
    print("HOME_COPY", HOME_COPY)
    print("SHEETS", " | ".join(sheets))
    print("METRICS", metric_count)
    print("MANDATORY", mandatory)


if __name__ == "__main__":
    main()
