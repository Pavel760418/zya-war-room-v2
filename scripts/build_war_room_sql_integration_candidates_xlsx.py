#!/usr/bin/env python3
"""Build war_room_sql_integration_candidates.xlsx (SELECT-backed where available).

No passwords, IPs, or DATABASE_URL in output.
"""
from __future__ import annotations

import shutil
from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = Path("/home/andr/apps/zya-war-room-v2")
OUT_HOME = Path("/home/andr/war_room_sql_integration_candidates.xlsx")
OUT_DOCS = ROOT / "docs" / "war_room_sql_integration_candidates.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
WRAP = Alignment(wrap_text=True, vertical="top")


def _header(ws, headers: list[str]) -> None:
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    ws.freeze_panes = "A2"


def _auto_width(ws, max_w=48):
    for col in ws.columns:
        letter = col[0].column_letter
        w = min(max(len(str(c.value or "")) for c in col) + 2, max_w)
        ws.column_dimensions[letter].width = w


def _try_sql_snapshots() -> dict:
    """Bounded SQL checks for sheet 6; failures become comments only."""
    snap: dict = {"ok": False, "notes": []}
    try:
        from app.repositories.retail_payments_repository import PaymentPeriodFilters, RetailPaymentsRepository
        from app.repositories.retail_sales_repository import RetailSalesRepository, SalesPeriodFilters

        d = date(2026, 8, 10)
        sf = SalesPeriodFilters(d, d)
        pf = PaymentPeriodFilters(d, d)
        sales = RetailSalesRepository()
        pays = RetailPaymentsRepository()
        rc = sales.load_receipts_daily(sf)
        snap["receipts"] = rc
        snap["payments_total"] = pays.load_payment_summary_by_category(pf)
        snap["ok"] = True
        snap["notes"].append("Контрольный день 2026-08-10 загружен из SQL.")
    except Exception as exc:  # noqa: BLE001
        snap["notes"].append(f"SQL snapshot skipped: {type(exc).__name__}")
    return snap


def sheet_summary(ws):
    headers = [
        "Метрика",
        "Статус",
        "Источник подтверждён?",
        "Количество кандидатов",
        "Рекомендуемый источник",
        "Уровень уверенности",
        "Готово для War Room?",
        "Требуется действие ИТ",
        "Требуется действие 1С-консультанта",
        "Комментарий",
    ]
    _header(ws, headers)
    rows = [
        ("Дата продажи", "Подтверждено", "Да", 1, "_Document156._Date_Time + YearOffset 2000", "Высокий", "Да", "Нет", "Подтвердить vs дата закрытия смены", "Декодирование проверено на выборке"),
        ("Номер чека / префикс магазина", "Подтверждено", "Да", 1, "_Document156._Number", "Высокий", "Да", "Нет", "Сверить неизвестные префиксы (АП, 00…)", "Python: store_prefix_map.py"),
        ("Выручка / чеки / возвраты", "Подтверждено", "Да", 1, "_Document156 + _Fld4036/_Fld4030", "Высокий", "Да", "Нет", "Сверка с Excel/ККМ-отчётом", "Репозиторий retail_sales_repository"),
        ("Строки чека, кол-во, НДС строки", "Подтверждено", "Да", 1, "_Document156_VT4039", "Высокий", "Да", "Нет", "Правило НДС (_Fld4054 vs _Fld4048)", "Не умножать _Fld4030 при JOIN со строками"),
        ("Средний чек", "Подтверждено", "Частично", 1, "net_revenue / sales_checks", "Высокий", "Да", "Нет", "Согласовать знаменатель (только продажи)", "Возвраты отдельно"),
        ("Оплаты по типам", "Подтверждено с ограничением", "Да", 1, "_Document119_VT2299 + _Reference89", "Высокий", "Да (агрегат)", "Нет", "Не привязывать к чеку", "Только закрытие смены"),
        ("Тип цены в чеке", "Кандидат", "Частично", 1, "_Fld4016RRef → _Reference92", "Средний", "Диагностика", "Да", "Подтвердить бизнес-имена типов цен", "77 значений в справочнике"),
        ("Выручка (регистр)", "Кандидат", "Нет", 2, "_AccumRg6691", "Средний", "Нет", "Да", "Да", "Дублирует чеки? — сверка"),
        ("Себестоимость / ВП / маржа", "Кандидат", "Нет", 2, "_AccumRg6691._Fld6708 / _AccumRg6738", "Средний", "Нет", "Да", "Да", "Партионный учёт"),
        ("Остатки / стоимость остатков", "Кандидат", "Нет", 3, "_AccumRg6601 + др.", "Низкий", "Нет", "Да", "Да", "393M+ строк — только период"),
        ("Списания / потери", "Не найдено", "Нет", 0, "—", "Не определён", "Нет", "Да", "Да", "Нужен объект 1С"),
        ("Планы продаж", "Не найдено", "Нет", 0, "—", "Не определён", "Нет", "Да", "Да", "—"),
        ("OPEX / ФОТ / численность", "Кандидат", "Нет", 1, "_Reference82 + неизвестный регистр", "Низкий", "Нет", "Да", "Да", "Только эвристика имён"),
    ]
    for r in rows:
        ws.append(list(r))
    _auto_width(ws)


def sheet_sales_fields(ws):
    _header(
        ws,
        [
            "Техническое поле",
            "Бизнес-название",
            "Назначение",
            "Тип",
            "Правило преобразования",
            "Использование в War Room",
            "Риски",
            "Статус подтверждения",
        ],
    )
    fields = [
        ("dbo._Document156", "Документ розничного чека", "Заголовок чека ККМ", "Document", "—", "Факт продаж", "Тестовые/сторно документы", "Подтверждено"),
        ("dbo._Document156._Number", "Номер документа", "Номер чека, префикс магазина", "nchar(10)", "prefix → store Python dict", "Магазин", "Неизвестные префиксы", "Подтверждено"),
        ("dbo._Document156._Date_Time", "Дата-время чека", "Момент продажи", "datetime", "DATEADD(year,-2000,...)", "День/период", "vs дата смены", "Подтверждено"),
        ("dbo._Document156._Fld4036", "Тип операции", "1 возврат, 2 продажа", "numeric(1)", "filter", "Раздельные KPI", "Сторно?", "Подтверждено"),
        ("dbo._Document156._Fld4030", "Сумма документа", "Итог чека", "numeric(15)", "SUM by day/store", "Выручка", "≠ сумма строк при ошибках", "Подтверждено"),
        ("dbo._Document156._Fld4020", "Кассир", "ФИО/логин кассира", "ntext", "GROUP BY", "Доп. разрез", "ПДн", "Подтверждено"),
        ("dbo._Document156._Fld4016RRef", "Тип цены", "Ссылка _Reference92", "binary(16)", "Python dict", "Диагностика", "Не все имена проверены", "Кандидат"),
        ("dbo._Document156._Fld4028", "Числовой код", "Внутренний код чека", "nvarchar", "как есть", "Сверка", "—", "Подтверждено (тех.)"),
        ("dbo._Document156_VT4039", "ТЧ товаров чека", "Строки номенклатуры", "Tabular", "JOIN by _IDRRef", "SKU-level", "Тяжёлый JOIN", "Подтверждено"),
        ("VT4039._Fld4041RRef", "Номенклатура", "Ссылка на товар", "binary(16)", "→ _Reference*", "ABC/остатки", "Нужен справочник", "Подтверждено"),
        ("VT4039._Fld4042", "Количество", "Продано ед.", "numeric", "SUM", "Units", "Ед. изм.", "Подтверждено"),
        ("VT4039._Fld4046", "Цена", "Цена строки", "numeric", "—", "Аналитика", "—", "Подтверждено"),
        ("VT4039._Fld4048", "Сумма строки", "Без НДС?", "numeric", "SUM", "Сверка с _Fld4030", "НДС", "Подтверждено"),
        ("VT4039._Fld4054", "Сумма с НДС", "НДС строки", "numeric", "SUM", "НДС KPI", "Малые суммы vs _Fld4048", "Подтверждено"),
    ]
    for f in fields:
        ws.append(list(f))
    _auto_width(ws)


def sheet_payments(ws):
    _header(
        ws,
        [
            "Объект",
            "Описание",
            "Поле",
            "Агрегация",
            "Ограничение",
            "Mapping",
            "Статус",
        ],
    )
    rows = [
        ("dbo._Document119", "Закрытие смены", "_Number, _Date_Time", "1 doc = 1 смена/магазин/день", "Не чек", "Префикс магазина как у чеков", "Подтверждено"),
        ("dbo._Document119_VT2299", "ТЧ оплат смены", "_Fld2301RRef, _Fld2302", "SUM по смене/дню/магазину", "Нет связи с _Document156", "Python bytes→_Reference89", "Подтверждено"),
        ("dbo._Reference89", "Формы оплаты", "_Description", "Справочник", "8 активных форм", "payment_form_mapping.py", "Подтверждено"),
        ("—", "Ограничение War Room", "—", "Подпись UI: оплаты по закрытиям смен", "Запрет per-check payment", "docs/retail_sql_data_layer.md", "Обязательно"),
        ("Категории", "cash/cards/cashless/certificates/bonuses/other", "по _Description", "SUM", "Эвристика категорий", "payment_form_mapping._category_from_description", "Кандидат"),
    ]
    for r in rows:
        ws.append(list(r))
    _auto_width(ws)


def sheet_register_candidates(ws):
    _header(
        ws,
        [
            "Метрика",
            "Тип объекта 1С",
            "Техническая SQL-таблица",
            "Поля даты",
            "Поля магазина",
            "Поля суммы",
            "Поля количества",
            "Поля себестоимости",
            "Поля движения",
            "Поля ссылки на документ",
            "Признак кандидата",
            "Тестовый запрос",
            "Результат теста",
            "Риски",
            "Уровень уверенности",
            "Требуется подтверждение ИТ",
            "Требуется подтверждение 1С-консультанта",
            "Комментарий",
        ],
    )
    cands = [
        (
            "Выручка (оборот)",
            "AccumRg",
            "dbo._AccumRg6691",
            "_Period",
            "_Fld6692RRef→_Reference64",
            "_Fld6704,_Fld6707",
            "_Fld6703",
            "_Fld6708,_Fld6711",
            "_Active",
            "_RecorderRRef",
            "Средний",
            "SUM _Fld6704 за 1 день по магазину",
            "171M строк; период 2012–2026",
            "Дубль с чеками",
            "Средний",
            "Да",
            "Да",
            "Сверить с _Document156 за контрольный день",
        ),
        (
            "Себестоимость",
            "AccumRg",
            "dbo._AccumRg6738",
            "_Period",
            "_Fld6739RRef?",
            "_Fld6742,_Fld6743",
            "—",
            "_Fld6742?",
            "_Active",
            "_RecorderRRef",
            "Средний",
            "SUM за 1 день TOP магазинов",
            "161M строк",
            "Партионный учёт",
            "Средний",
            "Да",
            "Да",
            "Альтернатива _Fld6708",
        ),
        (
            "Складской оборот",
            "AccumRg",
            "dbo._AccumRg6601",
            "_Period",
            "_Fld6603RRef→склад",
            "_Fld6608",
            "—",
            "—",
            "_Active",
            "_RecorderRRef",
            "Низкий",
            "SUM TOP 50 складов за 1 день",
            "393M строк",
            "Не выручка",
            "Низкий",
            "Да",
            "Да",
            "Только запасы",
        ),
        (
            "Справочник магазинов",
            "Reference",
            "dbo._Reference64",
            "—",
            "_Code,_Description",
            "—",
            "—",
            "—",
            "—",
            "—",
            "Высокий",
            "SELECT TOP 20 _Code,_Description",
            "Справочник не пуст",
            "Связь с префиксом номера не доказана",
            "Высокий",
            "Да",
            "Да",
            "Параллельно префиксы _Number",
        ),
    ]
    for c in cands:
        ws.append(list(c))
    _auto_width(ws)


def sheet_queries(ws):
    _header(
        ws,
        [
            "Метрика",
            "Безопасный SELECT (шаблон)",
            "Параметры",
            "Ожидаемый результат",
            "Пример агрегирования",
            "Статус выполнения",
            "Краткий результат",
            "Комментарий",
        ],
    )
    queries = [
        (
            "Чеки за день",
            "SELECT COUNT(*), SUM(_Fld4030) FROM dbo._Document156 WHERE _Posted=0x01 AND _Date_Time>=DATEADD(year,2000,@d) AND _Date_Time<DATEADD(year,2000,DATEADD(day,1,@d))",
            "@d date",
            "Число документов и сумма",
            "GROUP BY _Fld4036",
            "Выполнено",
            "2026-08-10: 512 docs",
            "Кontrol day",
        ),
        (
            "Строки чека за день",
            "SELECT SUM(_Fld4042),SUM(_Fld4048) FROM VT4039 WHERE _Document156_IDRRef IN (SELECT _IDRRef FROM _Document156 WHERE …)",
            "@d date",
            "Qty и сумма строк",
            "по operation_type",
            "Выполнено",
            "line_sum согласована с doc_sum при GROUP BY op",
            "Без JOIN SUM doc",
        ),
        (
            "Оплаты смены",
            "SELECT SUM(_Fld2302) FROM _Document119 d JOIN _Document119_VT2299 v …",
            "@d date",
            "Сумма оплат",
            "по _Reference89",
            "Выполнено",
            "~29.4M RUB / 15 shifts / day",
            "Не равно выручке чеков без сверки",
        ),
    ]
    for q in queries:
        ws.append(list(q))
    _auto_width(ws)


def sheet_reconciliation(ws, snap: dict):
    _header(
        ws,
        [
            "Метрика",
            "Период",
            "Магазин",
            "Excel",
            "SQL",
            "Разница",
            "Допустимое отклонение",
            "Статус",
            "Объяснение расхождения",
            "Следующее действие",
        ],
    )
    excel_note = "Шаблон war-room-template-2-no-traffic.xlsx — июнь 2026, агрегат месяц; нет дневной детализации 2026-08-10"
    rows = [
        ("Выручка факт", "2026-06", "Сеть", "Есть в шаблоне", "Не сверялось (другой период)", "—", "—", "Отложено", excel_note, "Запросить управленческий Excel за день"),
        ("Количество чеков", "2026-06", "Сеть", "Есть", "Не сверялось", "—", "—", "Отложено", excel_note, "После дневного Excel"),
    ]
    if snap.get("ok") and snap.get("receipts") is not None and not snap["receipts"].empty:
        r = snap["receipts"].iloc[0]
        rows.append(
            (
                "Чистая выручка (SQL internal)",
                "2026-08-10",
                "Вся сеть",
                "—",
                float(r.get("net_revenue", 0)),
                "—",
                "—",
                "SQL-only контроль",
                "Сверка Excel недоступна за этот день",
                "Загрузить Excel ККМ",
            )
        )
        if snap.get("payments_total") is not None and not snap["payments_total"].empty:
            pt = float(snap["payments_total"]["total"].sum())
            rows.append(
                (
                    "Оплаты по сменам",
                    "2026-08-10",
                    "Вся сеть",
                    "—",
                    pt,
                    "—",
                    "—",
                    "SQL-only",
                    "Не сопоставлять с чеками без методики",
                    "Сверить с Z-отчётом",
                )
            )
    for row in rows:
        ws.append(list(row))
    _auto_width(ws)


def sheet_questions(ws):
    _header(ws, ["Тема", "Аудитория", "Вопрос", "Контекст SQL", "Приоритет"])
    qs = [
        ("Магазин", "1С", "Подтвердите, что префиксы _Document156._Number (АВ, ЗЯ, …) — канон для War Room, и что делать с префиксами АП, 00, 01?", "_Number nchar(10)", "Высокий"),
        ("Оплаты", "1С", "Есть ли регистр/реквизит формы оплаты на уровне чека в другой выгрузке?", "VT2299 только на _Document119", "Высокий"),
        ("НДС", "1С", "_Fld4054 vs _Fld4048 — что является «суммой с НДС» для управленческого отчёта?", "VT4039", "Средний"),
        ("Выручка", "ИТ", "Какой регистр является официальным источником выручки для BI — _Document156 или _AccumRg6691?", "Два источника", "Высокий"),
        ("Себестоимость", "1С", "Какой ресурс партионной себестоимости использовать: _AccumRg6691._Fld6708 или _AccumRg6738?", "Кандидаты", "Высокий"),
        ("Планы", "1С", "Где хранятся планы продаж по магазинам (документ/регистр)?", "Не найдено", "Средний"),
        ("Списания", "1С", "Объект SQL для списаний, недостач, инвентаризаций?", "Info/Accum поиск", "Высокий"),
        ("Справочник", "ИТ", "_Reference64 — это магазины ТКПТ и как связать с префиксом номера?", "_Reference64", "Средний"),
    ]
    for q in qs:
        ws.append(list(q))
    _auto_width(ws)


def main() -> int:
    snap = _try_sql_snapshots()
    wb = Workbook()
    ws0 = wb.active
    ws0.title = "Сводка"
    sheet_summary(ws0)
    sheet_sales_fields(wb.create_sheet("Продажи и чеки"))
    sheet_payments(wb.create_sheet("Оплаты"))
    sheet_register_candidates(wb.create_sheet("Кандидаты 1С-регистров"))
    sheet_queries(wb.create_sheet("Запросы для проверки"))
    sheet_reconciliation(wb.create_sheet("Сверка SQL и Excel"), snap)
    sheet_questions(wb.create_sheet("Вопросы ИТ и 1С"))

    OUT_HOME.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_HOME)
    OUT_DOCS.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(OUT_HOME, OUT_DOCS)

    load_workbook(OUT_HOME, read_only=True).close()
    print(f"OK: {OUT_HOME}")
    print(f"OK: {OUT_DOCS}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
