"""1C storage-structure catalog → logical metadata name → physical MSSQL table.

Source of truth: ``data/catalog/StrukturaKhraneniiaBazyDannykh.xlsx`` (sheet TDSheet).
Columns: Метаданные | Имя таблицы хранения | Имя таблицы | Назначение.

Physical names in the workbook omit the leading underscore used by MSSQL
(``AccumRg6691`` → ``_AccumRg6691``). Tabular parts become ``_DocumentNNN_VTMMMM``.
"""
from __future__ import annotations

import re
from dataclasses import dataclass
from functools import lru_cache
from pathlib import Path
from typing import Iterable, Optional

from openpyxl import load_workbook

from app.core.config import BASE_DIR
from app.ingestion.text_utils import normalize

__all__ = [
    "CATALOG_PATH",
    "MetadataRow",
    "load_metadata_rows",
    "physical_table",
    "resolve_logical",
    "build_physical_map",
    "WAR_ROOM_LOGICAL_ALIASES",
]

CATALOG_PATH = BASE_DIR / "data" / "catalog" / "StrukturaKhraneniiaBazyDannykh.xlsx"

# Canonical War Room / catalog logical names → preferred TDSheet metadata keys.
WAR_ROOM_LOGICAL_ALIASES: dict[str, tuple[str, ...]] = {
    "РегистрНакопления.Продажи": ("РегистрНакопления.Продажи",),
    "ВыручкаИСебестоимостьПродаж": (
        "РегистрНакопления.Продажи",
        "ВыручкаИСебестоимостьПродаж",
    ),
    "ТоварыНаСкладах": (
        "РегистрНакопления.ОстаткиТоваровКомпании",
        "ТоварыНаСкладах",
    ),
    "РегистрНакопления.ОстаткиТоваровКомпании": ("РегистрНакопления.ОстаткиТоваровКомпании",),
    "РегистрНакопления.БюджетПродаж": ("РегистрНакопления.БюджетПродаж",),
    "Документ.БюджетПродаж": ("Документ.БюджетПродаж",),
    "Документ.БюджетНакладныхРасходовИДоходов": ("Документ.БюджетНакладныхРасходовИДоходов",),
    "Документ.СписаниеТоваров": ("Документ.СписаниеТоваров",),
    "Документ.Инвентаризация": ("Документ.Инвентаризация",),
    "Справочник.Номенклатура": ("Справочник.Номенклатура",),
    "Справочник.ПодразделенияКомпании": ("Справочник.ПодразделенияКомпании",),
    "Справочник.Магазины": ("Справочник.ПодразделенияКомпании", "Справочник.Магазины"),
    # Tabular parts (Имя таблицы хранения like Document107.VT1803)
    "Документ.БюджетПродаж.Товары": ("Document107.VT1803", "Документ.БюджетПродаж"),
    "Документ.БюджетНакладныхРасходовИДоходов.РасходыИДоходы": (
        "Document105.VT1724",
        "Документ.БюджетНакладныхРасходовИДоходов",
    ),
    "Документ.СписаниеТоваров.Товары": ("Document172.VT4675", "Документ.СписаниеТоваров"),
    "Документ.Инвентаризация.Товары": ("Document124.VT2532", "Документ.Инвентаризация"),
}


@dataclass(frozen=True)
class MetadataRow:
    metadata: str
    storage_name: str
    table_name: str
    purpose: str

    @property
    def physical(self) -> str:
        return storage_to_mssql(self.storage_name)


_VT_RE = re.compile(r"^(Document|AccumRg|Reference|InfoRg|DocumentJournal)(\d+)\.VT(\d+)$", re.I)
_PLAIN_RE = re.compile(r"^(Document|AccumRg|Reference|InfoRg|DocumentJournal|Enum)(\d+)$", re.I)


def storage_to_mssql(storage: str) -> str:
    """``AccumRg6691`` / ``Document107.VT1803`` → ``_AccumRg6691`` / ``_Document107_VT1803``."""
    s = (storage or "").strip()
    if not s:
        return ""
    if s.startswith("_"):
        return s
    m = _VT_RE.match(s)
    if m:
        return f"_{m.group(1)}{m.group(2)}_VT{m.group(3)}"
    m = _PLAIN_RE.match(s)
    if m:
        return f"_{m.group(1)}{m.group(2)}"
    # Already looks like physical fragment
    if re.match(r"^[A-Za-z]+\d+", s):
        return f"_{s.replace('.', '_')}"
    return s


@lru_cache(maxsize=1)
def load_metadata_rows(path: Optional[str] = None) -> tuple[MetadataRow, ...]:
    catalog = Path(path) if path else CATALOG_PATH
    if not catalog.is_file():
        return ()
    wb = load_workbook(catalog, read_only=True, data_only=True)
    try:
        ws = wb["TDSheet"] if "TDSheet" in wb.sheetnames else wb.active
        rows: list[MetadataRow] = []
        first = True
        for values in ws.iter_rows(values_only=True):
            if first:
                first = False
                # skip header
                continue
            meta = str(values[0] or "").strip()
            storage = str(values[1] or "").strip()
            table = str(values[2] or "").strip()
            purpose = str(values[3] or "").strip()
            if not meta and not storage:
                continue
            rows.append(MetadataRow(meta, storage, table, purpose))
        return tuple(rows)
    finally:
        wb.close()


def build_physical_map(
    *,
    purpose: Optional[str] = "Основная",
    path: Optional[str] = None,
) -> dict[str, str]:
    """Map metadata / table_name / storage → MSSQL physical name."""
    out: dict[str, str] = {}
    for row in load_metadata_rows(path):
        if purpose and row.purpose and row.purpose != purpose:
            # Keep VT / non-Основная if storage looks like VT
            if ".VT" not in row.storage_name and "VT" not in row.storage_name:
                continue
        phys = row.physical
        if not phys:
            continue
        for key in (row.metadata, row.table_name, row.storage_name, phys):
            if key:
                out[key] = phys
                out[normalize(key)] = phys
    return out


def _score_candidate(query: str, row: MetadataRow) -> float:
    qn = normalize(query)
    scores = [
        1.0 if normalize(row.metadata) == qn else 0.0,
        0.95 if normalize(row.table_name) == qn else 0.0,
        0.9 if normalize(row.storage_name) == qn else 0.0,
        0.85 if qn and qn in normalize(row.metadata) else 0.0,
        0.75 if qn and qn in normalize(row.table_name) else 0.0,
        0.7 if qn and normalize(row.metadata).endswith(qn.split(".")[-1] if "." in qn else qn) else 0.0,
    ]
    return max(scores)


def resolve_logical(
    logical: str,
    *,
    path: Optional[str] = None,
    prefer_purpose: str = "Основная",
) -> Optional[str]:
    """Resolve a logical 1C name to MSSQL physical table (with leading ``_``)."""
    aliases = WAR_ROOM_LOGICAL_ALIASES.get(logical, (logical,))
    rows = load_metadata_rows(path)
    if not rows:
        return None

    # Exact / alias pass on Основная first
    for alias in aliases:
        an = normalize(alias)
        for row in rows:
            if prefer_purpose and row.purpose != prefer_purpose and ".VT" not in row.storage_name:
                continue
            if normalize(row.metadata) == an or normalize(row.table_name) == an or normalize(row.storage_name) == an:
                return row.physical
            # Direct storage token like Document107.VT1803
            if alias.replace(" ", "") == row.storage_name.replace(" ", ""):
                return row.physical

    # Fuzzy: best score among aliases
    best: Optional[MetadataRow] = None
    best_score = 0.0
    for alias in aliases:
        for row in rows:
            if prefer_purpose and row.purpose not in ("", prefer_purpose) and ".VT" not in row.storage_name:
                continue
            sc = _score_candidate(alias, row)
            if sc > best_score:
                best_score = sc
                best = row
    if best is not None and best_score >= 0.7:
        return best.physical
    return None


def physical_table(logical: str, *, default: Optional[str] = None) -> str:
    """Convenience: resolve or return ``default`` / raise if missing."""
    found = resolve_logical(logical)
    if found:
        return found
    if default:
        return default
    raise KeyError(f"Нет физического маппинга для «{logical}» в StrukturaKhraneniiaBazyDannykh.xlsx")


def known_war_room_physicals() -> dict[str, str]:
    """Stable dict used by sql_extract / services."""
    keys = list(WAR_ROOM_LOGICAL_ALIASES.keys())
    out: dict[str, str] = {}
    for k in keys:
        try:
            out[k] = physical_table(k)
        except KeyError:
            continue
    # Hard fallbacks from confirmed TDSheet dump (if file missing in test stub)
    fallbacks = {
        "РегистрНакопления.Продажи": "_AccumRg6691",
        "ВыручкаИСебестоимостьПродаж": "_AccumRg6691",
        "ТоварыНаСкладах": "_AccumRg6601",
        "РегистрНакопления.ОстаткиТоваровКомпании": "_AccumRg6601",
        "РегистрНакопления.БюджетПродаж": "_AccumRg6450",
        "Документ.БюджетПродаж": "_Document107",
        "Документ.БюджетПродаж.Товары": "_Document107_VT1803",
        "Документ.БюджетНакладныхРасходовИДоходов": "_Document105",
        "Документ.БюджетНакладныхРасходовИДоходов.РасходыИДоходы": "_Document105_VT1724",
        "Документ.СписаниеТоваров": "_Document172",
        "Документ.СписаниеТоваров.Товары": "_Document172_VT4675",
        "Документ.Инвентаризация": "_Document124",
        "Документ.Инвентаризация.Товары": "_Document124_VT2532",
        "Справочник.Номенклатура": "_Reference58",
        "Справочник.ПодразделенияКомпании": "_Reference64",
        "Справочник.Магазины": "_Reference64",
    }
    for k, v in fallbacks.items():
        out.setdefault(k, v)
    return out


def iter_main_objects(path: Optional[str] = None) -> Iterable[MetadataRow]:
    for row in load_metadata_rows(path):
        if row.purpose == "Основная":
            yield row
